import openpyxl
import sys

# 엑셀 파일 읽기
try:
    wb = openpyxl.load_workbook('data/아이티 공사 현황.xlsx')
    ws = wb.active
    
    print("시트명:", wb.sheetnames)
    print("데이터 범위:", ws.calculate_dimension())
    print()
    
    print("첫 번째 행 (헤더):")
    header_row = []
    for cell in ws[1]:
        if cell.value:
            header_row.append(str(cell.value))
        else:
            header_row.append("")
    print("\t".join(header_row))
    print()
    
    print("처음 5행의 데이터:")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_data = [str(cell) if cell is not None else "" for cell in row]
        print(f"{i}행: {' | '.join(row_data)}")
        if i >= 5:
            break
            
    print(f"\n총 행 수: {ws.max_row}")
    print(f"총 열 수: {ws.max_column}")
    
except Exception as e:
    print(f"오류 발생: {e}")